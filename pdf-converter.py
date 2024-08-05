import openpyxl as pxl
from openpyxl.utils import get_column_letter
import os
from win32com.client import Dispatch

# Load the workbook
file_path = r'reformatted-output/pdf-june-billing.xlsx'  # Update with the path to your Excel file
wb = pxl.load_workbook(file_path)

# Function to set print settings for each sheet
def set_print_settings(sheet):
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0  # Set to 0 to fit all rows on one page

    # Calculate the number of columns and adjust width accordingly
    max_column = sheet.max_column
    for col in range(1, max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].bestFit = True

# Apply settings to all sheets
for sheet in wb.worksheets:
    set_print_settings(sheet)

# Save each sheet as a separate Excel file and convert to PDF
excel = Dispatch('Excel.Application')
excel.Visible = False

output_dir = 'pdf-file'  # Update with your desired output directory
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

for sheet_name in wb.sheetnames:
    temp_wb = pxl.Workbook()
    temp_sheet = temp_wb.active
    temp_sheet.title = sheet_name

    # Copy data from the original sheet to the new temporary sheet
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(values_only=True):
        temp_sheet.append(row)

    # Set the same print settings
    set_print_settings(temp_sheet)

    # Save the temporary workbook
    temp_excel_path = os.path.join(output_dir, f"{sheet_name}.xlsx")
    temp_wb.save(temp_excel_path)

    # Convert the temporary workbook to PDF
    temp_pdf_path = os.path.join(output_dir, f"{sheet_name}.pdf")
    wb = excel.Workbooks.Open(temp_excel_path)
    wb.ExportAsFixedFormat(0, temp_pdf_path)
    wb.Close(False)

excel.Quit()

print(f"All sheets have been converted to PDF and saved in {output_dir}.")
