from flask import Flask, request, send_file
from flask_cors import CORS
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

UPLOAD_FOLDER = 'uploads'
OUTPUT_FILE_PATH = 'transformed_data-2.xlsx'

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def extract_summary_data(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb['Summary']
    
    data = {
        "Customer Name": "Get On Infotech Pvt Ltd",
        "Start Date": ws['B7'].value,
        "End Date": ws['B8'].value,
        "Subscription ID": ws['B3'].value.split('/')[-1]
    }
    return data

def process_file(file_path, summary_data):
    df = pd.read_excel(file_path, sheet_name='Data')
    
    df.rename(columns={
        'Meter': 'Meter Name',
        'ServiceName': 'Service Type',
        'ResourceLocation': 'Region',
        'ResourceType': 'Resource Name',
        'Cost': 'Total Cost'
    }, inplace=True)

    df_transformed = df[['Meter Name', 'Service Type', 'Resource Name', 'Region', 'Total Cost']].copy()
    df_transformed.loc[:, 'Total Cost'] = df_transformed['Total Cost'].round(2)
    
    for col, val in summary_data.items():
        df_transformed[col] = val

    df_transformed = df_transformed[[
        "Customer Name", "Start Date", "End Date", "Subscription ID",
        "Meter Name", "Service Type", "Resource Name", "Region", "Total Cost"
    ]]

    total_cost_sum = df_transformed['Total Cost'].sum()
    total_cost_row = pd.DataFrame({
        "Customer Name": [''],
        "Start Date": [''],
        "End Date": [''],
        "Subscription ID": [''],
        "Meter Name": [''],
        "Service Type": [''],
        "Resource Name": [''],
        "Region": ['Total Cost'],
        "Total Cost": [total_cost_sum]
    })
    df_transformed = pd.concat([df_transformed, total_cost_row], ignore_index=True)
    
    return df_transformed

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        app.logger.error('No file part in the request')
        return 'No file part', 400

    file = request.files['file']
    if file.filename == '':
        app.logger.error('No selected file')
        return 'No selected file', 400

    if file:
        filename = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filename)
        
        summary_data = extract_summary_data(filename)
        df_transformed = process_file(filename, summary_data)
        sheet_name = os.path.basename(file.filename).replace('.xlsx', '')
        sheet_name = sheet_name[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_').replace('?', '_')
        
        with pd.ExcelWriter(OUTPUT_FILE_PATH, engine='openpyxl') as writer:
            df_transformed.to_excel(writer, sheet_name=sheet_name, index=False)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        return send_file(OUTPUT_FILE_PATH, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, port=5000)
