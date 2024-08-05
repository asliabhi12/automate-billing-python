import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Directory containing the input files
input_directory = 'input-file'  # Replace with your directory path

# Create a list of all Excel files in the directory
input_files = glob.glob(os.path.join(input_directory, '*.xlsx'))

# Function to process the "Summary" sheet and extract data
def extract_summary_data(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb['Summary']

    def format_date(date_value):
        if isinstance(date_value, datetime):
            return date_value.strftime('%B %d, %Y')
        return ''
    
    start_date = format_date(ws['C8'].value)
    end_date = format_date(ws['C9'].value)

    print(start_date)
    print(end_date)
    
    data = {
        "Customer Name": "",
        "Start Date": start_date,
        "End Date": end_date,
        "Subscription ID": ws['C5'].value.split('/')[-1]
    }
    return data

# Function to process each file
def process_file(file_path, summary_data):
    # Load the data sheet into a DataFrame
    df = pd.read_excel(file_path, sheet_name='Data')
    
    # Rename the columns to match the desired format
    df.rename(columns={
        'Meter': 'Meter Name',
        'ServiceName': 'Service Type',
        'ResourceLocation': 'Region',
        'ResourceType': 'Resource Name',
        'Cost': 'Total Cost'
    }, inplace=True)

    # Select the desired columns
    df_transformed = df[['Meter Name', 'Service Type', 'Resource Name', 'Region', 'Total Cost']].copy()

    # Round the Total Cost to two decimal places
    df_transformed.loc[:, 'Total Cost'] = df_transformed['Total Cost']
    
    # Add summary data to the DataFrame
    for col, val in summary_data.items():
        df_transformed[col] = val

    # Reorder columns to match the desired output
    df_transformed = df_transformed[[
        "Customer Name", "Start Date", "End Date", "Subscription ID",
        "Meter Name", "Service Type", "Resource Name", "Region", "Total Cost"
    ]]

    # Calculate the total cost and append it as the last row
    total_cost_sum = df_transformed['Total Cost'].sum()
    total_cost_row = pd.DataFrame({
        "Customer Name": [''],
        "Start Date": [''],
        "End Date": [''],
        "Subscription ID": [''],
        "Meter Name": [''],
        "Service Type": [''],
        "Resource Name": [''],
        "Region": ['Total Cost'],
        "Total Cost": [total_cost_sum]
    })
    df_transformed = pd.concat([df_transformed, total_cost_row], ignore_index=True)
    
    return df_transformed

# Create a Pandas Excel writer object
output_file_path = 'transformed_data-2.xlsx'  # Replace with your desired output file path
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    for file_path in input_files:
        # Extract summary data
        summary_data = extract_summary_data(file_path)
        
        # Process each file
        df_transformed = process_file(file_path, summary_data)
        
        # Get the sheet name from the input file name
        sheet_name = os.path.basename(file_path).replace('.xlsx', '')
        
        # Ensure the sheet name is valid (sheet names must be <= 31 characters and cannot contain certain characters)
        sheet_name = sheet_name[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_').replace('?', '_')
        
        # Write the transformed data to a new sheet
        df_transformed.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Access the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Autofit column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

print(f"Transformed data has been saved to {output_file_path}")
