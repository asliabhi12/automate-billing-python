import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.styles import Alignment

# Path to the existing transformed data
input_file_path = 'EXCEL.xlsx'  # Replace with your file path
output_file_path = 'reformatted_data-final.xlsx'  # Path for the new formatted file

# Load the transformed data
df = pd.read_excel(input_file_path, sheet_name=None)

# Prepare to save the reformatted data
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    for sheet_name, data in df.items():
        # Process each sheet
        df_transformed = data.copy()
        
        # Remove columns that are not needed in the new format
        columns_to_keep = ['Meter Name', 'Service Type', 'Resource Name', 'Region', 'Total Cost']
        df_filtered = df_transformed[columns_to_keep]

        # Remove unnecessary rows (total cost row)
        df_filtered = df_filtered[df_filtered['Meter Name'].notna()]

        # Add total cost row
        total_cost_sum = df_filtered['Total Cost'].sum()
        total_cost_row = pd.DataFrame({
            'Meter Name': [''],
            'Service Type': [''],
            'Resource Name': [''],
            'Region': ['Total Cost'],
            'Total Cost': [total_cost_sum]
        })
        final_df = pd.concat([df_filtered, total_cost_row], ignore_index=True)

        # Add subscription ID and formatted summary as a separate section
        subscription_id = df_transformed['Subscription ID'].iloc[0]
        # summary_section = pd.DataFrame({
        #     'Meter Name': [''],
        #     'Service Type': [''],
        #     'Resource Name': [''],
        #     'Region': ['Subscription ID:'],
        #     'Total Cost': [subscription_id]
        # })
        
        # Append summary section
        # final_df = pd.concat([df_filtered, summary_section], ignore_index=True)

        # Write the reformatted data to the new sheet
        final_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Access the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Autofit column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Merge cells for Subscription ID in the last row
        last_row = len(final_df)  # Last row with data
        num_columns = len(final_df.columns)  # Number of columns in the DataFrame

        # Optional: Adjust the alignment of the merged cell
        cell = worksheet.cell(row=last_row, column=1)
        cell.value = f"Subscription ID: {subscription_id}"
        cell.alignment = Alignment(horizontal='center')

print(f"Reformatted data has been saved to {output_file_path}")
